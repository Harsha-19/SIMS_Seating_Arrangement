import io
import qrcode  # type: ignore
from typing import cast, Any
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib import colors  # type: ignore
from reportlab.lib.pagesizes import A4  # type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak  # type: ignore
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore

from django.http import HttpResponse  # type: ignore
from api.utils import build_seat_grid

class ExcelExportService:
    @staticmethod
    def generate_plan_excel(plan):
        wb = Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)
        
        assignments = plan.assignments.select_related('student', 'room', 'exam_group', 'exam').all()
        rooms = list(set([a.room for a in assignments]))
        rooms.sort(key=lambda r: r.room_number)

        for room in rooms:
            sheet = cast(Worksheet, wb.create_sheet(title=f"Room {room.room_number}"))
            room_assignments = [a for a in assignments if a.room_id == room.id]
            
            # Header
            sheet.merge_cells('A1:E1')
            cell_a1 = cast(Cell, sheet.cell(row=1, column=1))
            cell_a1.value = f"Seating Plan: {plan.exam.subject} | Room: {room.room_number}"
            cell_a1.font = Font(bold=True, size=14)
            cell_a1.alignment = Alignment(horizontal='center')
            
            # Sub-header
            header_row: list[Any] = ["Row", "Column", "Seat Label", "Reg No", "Student Name", "Department", "Semester"]
            sheet.append(header_row)
            
            row_2 = sheet[2]
            if isinstance(row_2, tuple):
                for cell in row_2:
                    if hasattr(cell, 'font'):
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Data
            for a in sorted(room_assignments, key=lambda x: (x.row, x.seat_position)):
                row_data: list[Any] = [
                    a.row, 
                    a.seat_position.split('C')[-1] if 'C' in a.seat_position else '', 
                    a.seat_position, 
                    a.student.reg_no, 
                    a.student.name,
                    a.student.enrollments.first().program.name if a.student.enrollments.exists() else 'N/A',
                    a.student.enrollments.first().semester if a.student.enrollments.exists() else 'N/A'
                ]
                sheet.append(row_data)

            # Auto-column width
            for col in sheet.columns:
                max_length = 0
                if not col:
                    continue
                first_cell = col[0]
                col_idx = getattr(first_cell, 'column', None)
                if col_idx is None:
                    continue
                column_letter = get_column_letter(col_idx)
                for cell in col:
                    if hasattr(cell, 'value') and cell.value:
                        try:
                            val_len = len(str(cell.value))
                            if val_len > max_length:
                                max_length = val_len
                        except Exception:
                            pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

class PDFExportService:
    @staticmethod
    def _create_qr(data):
        qr = qrcode.QRCode(version=1, box_size=10, border=1)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        qr_io = io.BytesIO()
        cast(Any, img).save(qr_io, format='PNG')
        qr_io.seek(0)
        return qr_io

    @classmethod
    def generate_plan_pdf(cls, plan):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        assignments = plan.assignments.select_related('student', 'room', 'exam_group', 'exam').all()
        rooms = list(set([a.room for a in assignments]))
        rooms.sort(key=lambda r: r.room_number)

        for room in rooms:
            elements.append(Paragraph(f"<b>Seating Plan - {plan.exam.subject}</b>", styles['Title']))
            elements.append(Paragraph(f"Room: {room.room_number} | Date: {plan.exam.date}", styles['Normal']))
            elements.append(Spacer(1, 12))
            
            # Grid Visual
            grid = build_seat_grid(room)
            max_row = max([s['row'] for s in grid]) if grid else 0
            max_col = max([s['col'] for s in grid]) if grid else 0

            
            # Map assignments to grid
            room_map = {(a.row, a.seat_position): a for a in assignments if a.room_id == room.id}
            
            # Table data
            table_data = []
            for r in range(1, max_row + 1):
                row_data = []
                for c in range(1, max_col + 1):
                    label = f"R{r}C{c}"
                    # Find seat in grid to check if it's an aisle
                    is_aisle = not any(s['row'] == r and s['col'] == c for s in grid)
                    
                    if is_aisle:
                        row_data.append("")
                    else:
                        assign = None
                        for s_coords, s_assign in room_map.items():
                            if s_coords[0] == r and s_assign.seat_position == label:
                                assign = s_assign
                                break
                        
                        if assign:
                            content = f"{assign.student.name}\n{assign.student.reg_no}"
                        else:
                            content = "VACANT"
                        row_data.append(content)
                table_data.append(row_data)

            if table_data:
                t = Table(table_data, colWidths=[550/max_col]*max_col if max_col > 0 else [550])
                t.setStyle(TableStyle([
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('FONTSIZE', (0,0), (-1,-1), 6),
                ]))
                elements.append(t)
            
            elements.append(PageBreak())

            # Attendance Sheet for this room
            elements.append(Paragraph(f"<b>Attendance Sheet - {room.room_number}</b>", styles['Title']))
            room_assignments = sorted([a for a in assignments if a.room_id == room.id], key=lambda x: x.student.reg_no)
            
            attend_data: list[list[Any]] = [["SL", "Reg No", "Student Name", "Seat", "Signature"]]
            for i, a in enumerate(room_assignments):
                attend_data.append([str(i+1), a.student.reg_no, a.student.name, a.seat_position, " [ ] "])
            
            at = Table(attend_data, colWidths=[30, 80, 200, 80, 100])
            at.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('FONTSIZE', (0,0), (-1,-1), 9),
            ]))
            elements.append(at)
            elements.append(PageBreak())

        doc.build(elements)
        buffer.seek(0)
        return buffer

    @classmethod
    def generate_attendance_pdf(cls, plan):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        assignments = plan.assignments.select_related('student', 'room', 'exam').order_by('room__room_number', 'student__reg_no')
        rooms = list(set([a.room for a in assignments]))
        rooms.sort(key=lambda r: r.room_number)

        for room in rooms:
            elements.append(Paragraph(f"<b>Attendance Sheet - {room.room_number}</b>", styles['Title']))
            elements.append(Paragraph(f"Exam: {plan.exam.subject} | Date: {plan.exam.date}", styles['Normal']))
            elements.append(Spacer(1, 12))
            
            room_assignments = [a for a in assignments if a.room_id == room.id]
            attend_data: list[list[Any]] = [["SL", "Reg No", "Student Name", "Seat", "Signature"]]
            for i, a in enumerate(room_assignments):
                attend_data.append([str(i+1), a.student.reg_no, a.student.name, a.seat_position, " ____________ "])
            
            at = Table(attend_data, colWidths=[30, 80, 200, 80, 100])
            at.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('ALIGN', (0,0), (0,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(at)
            elements.append(PageBreak())

        doc.build(elements)
        buffer.seek(0)
        return buffer

    @classmethod
    def generate_hall_tickets(cls, plan):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        ticket_style = ParagraphStyle('Ticket', parent=styles['Normal'], fontSize=8)

        assignments = plan.assignments.select_related('student', 'room', 'exam').order_by('student__reg_no')

        # 4 tickets per page
        for i, a in enumerate(assignments):
            qr_data = f"Plan:{plan.id}|Student:{a.student.reg_no}|Seat:{a.seat_position}"
            qr_io = cls._create_qr(qr_data)
            qr_img = Image(qr_io, width=80, height=80)

            # Ticket Content
            details = [
                [Paragraph(f"<b>HALL TICKET</b><br/>{plan.exam.subject}", styles['Normal']), qr_img],
                [Paragraph(f"Student: {a.student.name}<br/>Reg: {a.student.reg_no}", ticket_style), ""],
                [Paragraph(f"Room: {a.room.room_number}<br/>Seat: {a.seat_position}", ticket_style), ""]
            ]
            
            t = Table(details, colWidths=[200, 100])
            t.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 1, colors.black),
                ('SPAN', (1,0), (1,2)),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            
            elements.append(t)
            elements.append(Spacer(1, 24))
            
            if (i + 1) % 4 == 0:
                elements.append(PageBreak())
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
