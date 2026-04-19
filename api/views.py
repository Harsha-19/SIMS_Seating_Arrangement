from rest_framework import viewsets, status, generics
from rest_framework.response import Response
from rest_framework.decorators import action, api_view
from rest_framework.exceptions import ValidationError as DRFValidationError
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import AllowAny
from django.db import transaction, models
from django.db.models import Prefetch
from django.core.exceptions import ValidationError as DjangoValidationError
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from rest_framework_simplejwt.tokens import RefreshToken
from .models import (
    Student, Room, Exam, ExamGroup, Seating, Attendance, AttendanceEntry, Program, Enrollment,
    ODD_SEMESTERS, EVEN_SEMESTERS, VALID_SEMESTERS, derive_sem_type,
)
from .serializers import (
    StudentSerializer, RoomSerializer, ExamSerializer, 
    SeatingSerializer, AttendanceSerializer, UserSerializer,
    ProgramSerializer, EnrollmentSerializer, SemesterOptionSerializer
)
from .utils import (
    sanitize_academic_info,
    extract_semester_info,
    build_seating_candidate,
    extract_student_identifier,
    orchestrate_exam_seating,
    sort_students_by_university_id,
    students_are_sorted_by_university_id,
)
import logging
from collections.abc import Mapping

logger = logging.getLogger(__name__)


def _snapshot_request_data(payload):
    if isinstance(payload, Mapping):
        if hasattr(payload, 'lists'):
            snapshot = {}
            for key, values in payload.lists():
                snapshot[key] = values if len(values) != 1 else values[0]
            return snapshot
        return {key: payload.get(key) for key in payload.keys()}
    return payload


def get_allowed_semesters(semester_type):
    if semester_type == 'ODD':
        return [1, 3, 5]
    if semester_type == 'EVEN':
        return [2, 4, 6]
    raise ValueError('Invalid semester type')


def normalize_semester_type(semester_type, *, required=False):
    if semester_type is None or str(semester_type).strip() == '':
        if required:
            raise ValueError('semester_type is required and must be either "ODD" or "EVEN".')
        return ''

    normalized = str(semester_type).strip().upper()
    get_allowed_semesters(normalized)
    return normalized


def _response_errors(message, errors=None, **extra):
    payload = {'message': message, 'error': message}
    if errors:
        payload['errors'] = errors
    payload.update(extra)
    return payload


def _parse_generate_seating_payload(payload):
    if not isinstance(payload, Mapping):
        raise ValueError('Expected a JSON object payload.')

    exam_id = payload.get('exam_id')
    room_ids = payload.get('rooms', payload.get('room_ids'))
    semester_type = payload.get('semester_type')

    errors = {}

    if exam_id in (None, ''):
        errors['exam_id'] = ['This field is required.']

    if room_ids in (None, ''):
        errors['rooms'] = ['This field is required and must be a non-empty array of room IDs.']
    elif not isinstance(room_ids, list) or not room_ids:
        errors['rooms'] = ['This field must be a non-empty array of room IDs.']

    if errors:
        raise ValueError(_response_errors('Invalid seating generation payload.', errors))

    try:
        parsed_exam_id = int(exam_id)
    except (TypeError, ValueError) as exc:
        raise ValueError(_response_errors('Invalid seating generation payload.', {
            'exam_id': ['exam_id must be an integer.'],
        })) from exc

    try:
        parsed_room_ids = [int(room_id) for room_id in room_ids]
    except (TypeError, ValueError) as exc:
        raise ValueError(_response_errors('Invalid seating generation payload.', {
            'rooms': ['Each room ID must be an integer.'],
        })) from exc

    normalized_semester_type = normalize_semester_type(semester_type)

    return parsed_exam_id, parsed_room_ids, normalized_semester_type


def _resolve_exam_groups(exam):
    groups = list(exam.groups.select_related('program').order_by('program__name', 'semester', 'subject', 'id'))
    if groups:
        return groups

    if exam.program and exam.semester:
        return [ExamGroup(exam=exam, program=exam.program, semester=exam.semester, subject=exam.subject)]

    return []


def _build_grouped_candidates(exam_groups):
    grouped_candidates = []
    matched_students = 0
    group_logs = []
    seen_student_groups = {}
    student_identifiers = {}

    for group in exam_groups:
        enrollments = (
            Enrollment.objects
            .filter(program=group.program, semester=group.semester)
            .select_related('student', 'program')
            .order_by('student__reg_no')
        )
        ordered_students = sort_students_by_university_id(enrollment.student for enrollment in enrollments)

        if not students_are_sorted_by_university_id(ordered_students):
            raise AssertionError("Students are not sorted properly")

        candidate_group = []
        for student in ordered_students:
            seen_student_groups.setdefault(student.id, []).append(group)
            student_identifiers[student.id] = extract_student_identifier(student)
            candidate_group.append(
                build_seating_candidate(
                    student,
                    exam_group=group if getattr(group, 'pk', None) else None,
                    subject=group.subject,
                    program_name=group.program.name,
                    semester=group.semester,
                    group_key=(group.program.name, group.semester, group.subject),
                )
            )

        grouped_candidates.append(candidate_group)
        matched_students += len(candidate_group)
        group_logs.append(
            f"{group.program.name} SEM {group.semester} - {group.subject}: {len(candidate_group)} student(s)"
        )

    duplicate_students = {
        student_id: groups
        for student_id, groups in seen_student_groups.items()
        if len(groups) > 1
    }

    return grouped_candidates, matched_students, group_logs, duplicate_students, student_identifiers


def _build_unsafe_seating_response(result, exam):
    diagnostics = result['diagnostics']
    errors = {}

    if diagnostics['consecutive_id_adjacent']:
        errors['university_id'] = [
            f"Unsafe layout: {diagnostics['consecutive_id_adjacent']} adjacent consecutive university-id pair(s) remain."
        ]

    if exam.exam_type == Exam.CORE and diagnostics['subject_count'] > 1 and diagnostics['same_subject_adjacent']:
        errors['subject'] = [
            f"Unsafe layout: {diagnostics['same_subject_adjacent']} adjacent same-subject pair(s) remain."
        ]

    if not errors:
        return None

    return Response(
        _response_errors(
            'Unsafe seating layout. Add more rooms or split the exam groups further.',
            errors,
            diagnostics=diagnostics,
        ),
        status=status.HTTP_400_BAD_REQUEST,
    )


def _build_generate_seating_response(request):
    payload = request.data if isinstance(request.data, Mapping) else {}
    request_snapshot = {
        'exam_id': payload.get('exam_id'),
        'rooms': payload.get('rooms', payload.get('room_ids')),
        'semester_type': payload.get('semester_type'),
    }
    logger.info("generate_seating request received: %s", request_snapshot)

    exam_id, room_ids, semester_type = _parse_generate_seating_payload(payload)

    exam = (
        Exam.objects
        .select_related('program')
        .prefetch_related('groups__program')
        .filter(id=exam_id)
        .first()
    )
    if not exam:
        logger.warning("generate_seating failed: invalid exam_id=%s", exam_id)
        return Response(
            _response_errors('Invalid exam', details={'exam_id': exam_id}),
            status=status.HTTP_404_NOT_FOUND,
        )

    exam_groups = _resolve_exam_groups(exam)
    if not exam_groups:
        logger.warning("generate_seating failed: exam_id=%s has no exam groups", exam.id)
        return Response(
            _response_errors(
                'Exam has no configured groups.',
                {'groups': ['Add at least one exam group before generating seating.']},
            ),
            status=status.HTTP_400_BAD_REQUEST,
        )

    logger.info(
        "generate_seating exam lookup: exam_id=%s, exam_type=%s, group_count=%s, semesters=%s",
        exam.id,
        exam.exam_type,
        len(exam_groups),
        [group.semester for group in exam_groups],
    )

    if semester_type and len(exam_groups) == 1:
        allowed_semesters = get_allowed_semesters(semester_type)
        if exam_groups[0].semester not in allowed_semesters:
            logger.warning(
                "generate_seating failed: semester_type=%s does not match exam semester=%s",
                semester_type,
                exam_groups[0].semester,
            )
            return Response(
                _response_errors(
                    f'Exam semester {exam_groups[0].semester} does not belong to the {semester_type} semester set {list(allowed_semesters)}.',
                    {
                        'semester_type': [f'{semester_type} is not valid for exam semester {exam_groups[0].semester}.'],
                    },
                ),
                status=status.HTTP_400_BAD_REQUEST,
            )

    rooms = list(Room.objects.filter(id__in=room_ids).order_by('room_number'))
    if len(rooms) != len(set(room_ids)):
        logger.warning("generate_seating failed: invalid room ids=%s", room_ids)
        return Response(
            _response_errors(
                'One or more selected rooms were not found.',
                {'rooms': ['One or more room IDs are invalid.']},
            ),
            status=status.HTTP_400_BAD_REQUEST,
        )

    grouped_candidates, matched_students, group_logs, duplicate_students, student_identifiers = _build_grouped_candidates(exam_groups)
    total_capacity = sum(room.total_capacity for room in rooms)
    logger.info(
        "generate_seating filters resolved: exam_id=%s, exam_type=%s, rooms=%s, group_logs=%s, matched_students=%s, capacity=%s",
        exam.id,
        exam.exam_type,
        [room.id for room in rooms],
        group_logs,
        matched_students,
        total_capacity,
    )

    if duplicate_students:
        duplicate_messages = []
        for student_id, groups in duplicate_students.items():
            group_labels = ', '.join(
                f"{group.program.name} SEM {group.semester} - {group.subject}"
                for group in groups
            )
            duplicate_messages.append(
                f"Student {student_identifiers.get(student_id, student_id)} matched multiple groups: {group_labels}."
            )
        logger.warning("generate_seating failed: duplicate student matches for exam_id=%s", exam.id)
        return Response(
            _response_errors(
                'Some students match more than one exam group.',
                {'groups': duplicate_messages},
            ),
            status=status.HTTP_400_BAD_REQUEST,
        )

    if matched_students == 0:
        logger.warning(
            "generate_seating failed: no enrollments found for exam_id=%s",
            exam.id,
        )
        return Response(
            _response_errors(
                'No students matched the configured exam groups.',
                {
                    'groups': ['No enrollments match the current exam groups.'],
                },
            ),
            status=status.HTTP_404_NOT_FOUND,
        )

    if matched_students > total_capacity:
        logger.warning(
            "generate_seating failed: insufficient capacity for exam_id=%s, matched_students=%s, capacity=%s",
            exam.id,
            matched_students,
            total_capacity,
        )
        return Response(
            _response_errors(
                f'Insufficient room capacity. Need {matched_students} seats but only {total_capacity} are available.',
                {'rooms': ['Selected rooms do not have enough capacity for all matched students.']},
            ),
            status=status.HTTP_400_BAD_REQUEST,
        )

    result = orchestrate_exam_seating(grouped_candidates, rooms, exam.exam_type, split_single_group=True)
    ordered_ids = result['ordered_ids']

    logger.info(
        "generate_seating ordered student ids: first=%s last=%s count=%s",
        ordered_ids[0] if ordered_ids else None,
        ordered_ids[-1] if ordered_ids else None,
        len(ordered_ids),
    )

    logger.info(
        "generate_seating allocation completed: allocated=%s, remaining=%s, diagnostics=%s",
        result['metrics']['allocated'],
        result['metrics']['remaining'],
        result['diagnostics'],
    )

    unsafe_response = _build_unsafe_seating_response(result, exam)
    if unsafe_response is not None:
        logger.warning("generate_seating failed safety validation for exam_id=%s", exam.id)
        return unsafe_response

    with transaction.atomic():
        Seating.objects.filter(exam=exam).delete()
        for seat in result['assignments']:
            Seating.objects.create(
                exam=exam,
                exam_group=seat['exam_group'],
                room=seat['room'],
                row=seat['row'],
                seat_position=seat['seat_pos'],
                student=seat['student'],
            )

    return Response(
        {
            'message': f'Generated {len(result["assignments"])} seats.',
            'exam_id': exam.id,
            'exam_type': exam.exam_type,
            'semester_type': semester_type or '',
            'matched_students': matched_students,
            'metrics': result['metrics'],
            'diagnostics': result['diagnostics'],
            'logs': [
                f"Loaded exam {exam.subject} ({exam.exam_type}) with {len(exam_groups)} group(s).",
                f"Validated {len(rooms)} room(s) with total capacity {total_capacity}.",
                *group_logs,
                f"Matched {matched_students} enrollment(s) across the exam groups.",
                f"Ordered students from {ordered_ids[0]} to {ordered_ids[-1]}." if ordered_ids else "No students matched.",
                f"Safety diagnostics: same_subject_adjacent={result['diagnostics']['same_subject_adjacent']}, consecutive_id_adjacent={result['diagnostics']['consecutive_id_adjacent']}.",
                *result['logs'],
                f"Placed {result['metrics']['allocated']} student(s).",
            ],
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
def generate_seating(request):
    try:
        return _build_generate_seating_response(request)
    except ValueError as exc:
        error_payload = exc.args[0] if exc.args and isinstance(exc.args[0], dict) else _response_errors(str(exc))
        logger.warning("generate_seating validation failed: %s", error_payload)
        return Response(error_payload, status=status.HTTP_400_BAD_REQUEST)
    except Exception:
        logger.exception("generate_seating failed unexpectedly")
        return Response(
            _response_errors('Unable to generate seating due to an internal server error.'),
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(['GET'])
def semester_options(request):
    department_id = request.query_params.get('department_id') or request.query_params.get('program_id')
    if not department_id:
        return Response({'error': 'department_id required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        department_id = int(department_id)
    except (TypeError, ValueError):
        return Response({'error': 'department_id must be an integer'}, status=status.HTTP_400_BAD_REQUEST)

    if not Program.objects.filter(id=department_id).exists():
        return Response({'error': 'Department not found'}, status=status.HTTP_404_NOT_FOUND)

    semester_payload = [
        {
            'id': semester,
            'number': semester,
            'name': f'Sem {semester}',
            'department': department_id,
        }
        for semester in VALID_SEMESTERS
    ]
    serializer = SemesterOptionSerializer(semester_payload, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all().prefetch_related('enrollments__program').order_by('-created_at')
    serializer_class = StudentSerializer
    parser_classes = (JSONParser, MultiPartParser, FormParser)
    lookup_field = 'reg_no'

    def _normalize_semester_type(self, semester_type):
        return normalize_semester_type(semester_type)

    def _build_enrollment_queryset(self, semester_type=None):
        if semester_type == 'ODD':
            queryset = Enrollment.objects.filter(
                sem_type='ODD',
                semester__in=ODD_SEMESTERS,
            )
        elif semester_type == 'EVEN':
            queryset = Enrollment.objects.filter(
                sem_type='EVEN',
                semester__in=EVEN_SEMESTERS,
            )
        elif semester_type:
            queryset = Enrollment.objects.none()
        else:
            queryset = Enrollment.objects.all()

        return queryset.select_related('student', 'program').order_by(
            'student_id', 'program__name', 'semester', 'section', '-created_at'
        )

    def _build_queryset(self, semester_type=None):
        enrollment_queryset = self._build_enrollment_queryset(semester_type)
        program_id = self.request.query_params.get('program_id')
        semester = self.request.query_params.get('semester')
        section = self.request.query_params.get('section')
        search = self.request.query_params.get('search')

        if program_id:
            enrollment_queryset = enrollment_queryset.filter(program_id=program_id)
        if semester:
            try:
                semester = int(semester)
            except (TypeError, ValueError):
                raise DjangoValidationError("semester must be an integer between 1 and 6.")
            if semester not in VALID_SEMESTERS:
                raise DjangoValidationError("semester must be between 1 and 6.")
            enrollment_queryset = enrollment_queryset.filter(semester=semester)
        if section:
            enrollment_queryset = enrollment_queryset.filter(section=section)

        if search:
            enrollment_queryset = enrollment_queryset.filter(
                models.Q(student__name__icontains=search) |
                models.Q(student__reg_no__icontains=search)
            )

        matching_enrollments = enrollment_queryset.filter(student_id=models.OuterRef('pk'))
        queryset = (
            Student.objects
            .annotate(has_matching_enrollment=models.Exists(matching_enrollments))
            .filter(has_matching_enrollment=True)
            .prefetch_related(
                Prefetch(
                    'enrollments',
                    queryset=enrollment_queryset,
                    to_attr='filtered_enrollments',
                )
            )
            .order_by('-created_at')
        )
        return queryset, enrollment_queryset.distinct()

    def get_queryset(self):
        semester_type = self.request.query_params.get('semester_type') or self.request.query_params.get('sem_type')
        normalized_semester_type = self._normalize_semester_type(semester_type)
        queryset, _ = self._build_queryset(normalized_semester_type)
        return queryset

    def list(self, request, *args, **kwargs):
        try:
            semester_type = request.GET.get("semester_type", None)
            semester_type = self._normalize_semester_type(
                semester_type or request.query_params.get('sem_type')
            )
            queryset, enrollment_queryset = self._build_queryset(semester_type)
            odd_students, odd_enrollments = self._build_queryset('ODD')
            even_students, even_enrollments = self._build_queryset('EVEN')
            odd_count = odd_students.count()
            even_count = even_students.count()
            filtered_count = queryset.count()
            odd_enrollment_ids = set(odd_enrollments.values_list('id', flat=True))
            even_enrollment_ids = set(even_enrollments.values_list('id', flat=True))
            overlap_count = len(odd_enrollment_ids & even_enrollment_ids)
            invalid_enrollment_count = Enrollment.objects.filter(
                models.Q(semester__isnull=True) |
                ~models.Q(semester__in=VALID_SEMESTERS)
            ).count()
            sem_type_mismatch_count = Enrollment.objects.filter(
                models.Q(semester__in=ODD_SEMESTERS, sem_type='EVEN') |
                models.Q(semester__in=EVEN_SEMESTERS, sem_type='ODD')
            ).count()
            warnings = []

            logger.debug("Student list semester_type=%s", semester_type or "ALL")
            logger.debug("Student list SQL=%s", enrollment_queryset.query)
            logger.debug(
                "Student list counts: filtered=%s enrollment_rows=%s odd=%s even=%s semesters=%s",
                filtered_count,
                enrollment_queryset.count(),
                odd_count,
                even_count,
                list(enrollment_queryset.values_list('semester', flat=True)[:20]),
            )
            logger.debug("Student list ODD SQL=%s", odd_enrollments.query)
            logger.debug("Student list EVEN SQL=%s", even_enrollments.query)
            if overlap_count:
                overlap_warning = f"ODD and EVEN enrollment querysets overlap on {overlap_count} record(s)."
                logger.warning(overlap_warning)
                warnings.append(overlap_warning)
            if invalid_enrollment_count:
                logger.warning("Found %s invalid enrollment semester records.", invalid_enrollment_count)
                warnings.append(f'Found {invalid_enrollment_count} invalid enrollment semester records.')
            if sem_type_mismatch_count:
                logger.warning("Found %s semester/sem_type mismatch records.", sem_type_mismatch_count)
                warnings.append(f'Found {sem_type_mismatch_count} semester/sem_type mismatch records.')

            serializer = self.get_serializer(queryset, many=True)
            return Response({
                'results': serializer.data,
                'counts': {
                    'filtered': filtered_count,
                    'odd': odd_count,
                    'even': even_count,
                },
                'semester_type': semester_type or 'ALL',
                'warning': ' '.join(warnings),
                'invalid_enrollment_count': invalid_enrollment_count,
                'diagnostics': {
                    'odd_even_overlap': overlap_count,
                    'semester_type_mismatch_count': sem_type_mismatch_count,
                },
            })
        except (DjangoValidationError, ValueError) as exc:
            if isinstance(exc, ValueError):
                detail = [str(exc)]
            else:
                detail = exc.message_dict if hasattr(exc, 'message_dict') else exc.messages
            return Response({'message': detail}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def upload(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'message': 'Excel file required.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_obj, read_only=True)
            sheet = wb.active
            
            header_map = {
                'reg_no': ['usn', 'roll', 'university', 'registration', 'reg', 'id'],
                'name': ['name', 'full name', 'student'],
                'class': ['class', 'semester', 'sem', 'course']
            }
            
            header_idx = {}
            header_row_num = 0
            for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                if not any(row): continue
                low_row = [str(cell).lower().strip() if cell else "" for cell in row]
                temp_map = {}
                matches = 0
                for logic_key, variants in header_map.items():
                    for i, h in enumerate(low_row):
                        if any(v in h for v in variants):
                            temp_map[logic_key] = i
                            matches += 1
                            break
                if matches >= 2:
                    header_idx = temp_map
                    header_row_num = row_num
                    break
            
            if len(header_idx) < 2:
                return Response({'message': 'Required columns missing.'}, status=status.HTTP_400_BAD_REQUEST)

            idx_reg = header_idx.get('reg_no')
            idx_name = header_idx.get('name')
            idx_class = header_idx.get('class')
            
            parsed_rows = []
            total_rows = 0
            
            for row in sheet.iter_rows(min_row=header_row_num + 1, values_only=True):
                if not any(row): continue
                get_val = lambda idx: str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else ""
                
                reg_no = get_val(idx_reg).upper()
                name = get_val(idx_name).upper()
                class_str = get_val(idx_class)
                
                if not class_str: class_str = " ".join([str(c) for c in row if c])
                academic = sanitize_academic_info(class_str)
                if not academic or not reg_no or not name: continue
                
                total_rows += 1
                parsed_rows.append({'reg_no': reg_no, 'name': name, 'academic': academic})

            new_records = 0
            duplicates_skipped = 0
            odd_count, even_count = 0, 0
            
            with transaction.atomic():
                program_cache = {p.name: p for p in Program.objects.all()}
                
                for r_data in parsed_rows:
                    # 1. Get/Create Student
                    student, _ = Student.objects.get_or_create(
                        reg_no=r_data['reg_no'],
                        defaults={"name": r_data['name']}
                    )
                    
                    # 2. Get Program (Strict match from Master Table)
                    p_name = r_data['academic']['program']
                    if p_name not in program_cache:
                        # Fallback for misspelled/new variants if master data is messy
                        program, _ = Program.objects.get_or_create(name=p_name)
                        program_cache[p_name] = program
                    program = program_cache[p_name]
                    
                    # 3. Create Enrollment safely (Multi-semester support)
                    enrollment, created = Enrollment.objects.get_or_create(
                        student=student, 
                        program=program, 
                        semester=r_data['academic']['semester'],
                        defaults={
                            "section": r_data['academic']['section'], 
                            "sem_type": derive_sem_type(r_data['academic']['semester'])
                        }
                    )
                    
                    if created:
                        new_records += 1
                    else:
                        duplicates_skipped += 1
                    
                    if enrollment.semester in ODD_SEMESTERS:
                        odd_count += 1
                    elif enrollment.semester in EVEN_SEMESTERS:
                        even_count += 1

            return Response({
                "total": total_rows, 
                "inserted": new_records, 
                "existing": duplicates_skipped, 
                "odd": odd_count, 
                "even": even_count
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'message': f'Upload failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            return Response({'message': f'Upload failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['delete'], url_path='bulk-delete')
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'message': 'No IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        Student.objects.filter(reg_no__in=ids).delete()
        return Response({'message': f'Deleted {len(ids)} students'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['delete'], url_path='clear-all')
    def clear_all(self, request):
        with transaction.atomic():
            Enrollment.objects.all().delete()
            Student.objects.all().delete()
        return Response({'message': 'All student records cleared'}, status=status.HTTP_200_OK)

class ProgramViewSet(viewsets.ModelViewSet):
    queryset = Program.objects.all().order_by('name')
    serializer_class = ProgramSerializer


class DepartmentViewSet(ProgramViewSet):
    """
    Backward-compatible alias for legacy frontend code that still uses
    `/api/departments/` after the Department model was renamed to Program.
    """
    queryset = Program.objects.all().order_by('-created_at')

    def _validate_department_payload(self, request, *, partial=False, instance=None):
        print("Incoming Data:", request.data)

        if not isinstance(request.data, Mapping):
            errors = {'non_field_errors': ['Expected a JSON object like {"name": "BCA"} .']}
            print("Serializer Errors:", errors)
            return None, Response(errors, status=status.HTTP_400_BAD_REQUEST)

        if 'name' not in request.data and 'department' in request.data:
            errors = {'name': ['Use "name" instead of "department".']}
            print("Serializer Errors:", errors)
            return None, Response(errors, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(instance=instance, data=request.data, partial=partial)
        if serializer.is_valid():
            return serializer, None

        print("Serializer Errors:", serializer.errors)
        logger.warning(
            "Department create validation failed",
            extra={
                'request_data': dict(request.data),
                'errors': serializer.errors,
            },
        )
        return None, Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def create(self, request, *args, **kwargs):
        serializer, error_response = self._validate_department_payload(request)
        if error_response:
            return error_response

        instance = serializer.save()
        response_serializer = self.get_serializer(instance)
        return Response(
            {
                'message': 'Department created',
                'department': response_serializer.data,
            },
            status=status.HTTP_201_CREATED,
        )

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer, error_response = self._validate_department_payload(
            request,
            partial=True,
            instance=instance,
        )
        if error_response:
            return error_response

        instance = serializer.save()
        response_serializer = self.get_serializer(instance)
        return Response(
            {
                'message': 'Department updated',
                'department': response_serializer.data,
            },
            status=status.HTTP_200_OK,
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        name = instance.name
        instance.delete()
        return Response(
            {'message': f"Department '{name}' deleted."},
            status=status.HTTP_200_OK,
        )

class EnrollmentViewSet(viewsets.ModelViewSet):
    queryset = Enrollment.objects.all().select_related('student', 'program')
    serializer_class = EnrollmentSerializer

class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.all().order_by('room_number')
    serializer_class = RoomSerializer

class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.all().select_related('program').prefetch_related('groups__program').order_by('date', 'start_time')
    serializer_class = ExamSerializer

    def _save_exam(self, request, *, partial=False, instance=None):
        action_name = 'create'
        if instance is not None:
            action_name = 'partial_update' if partial else 'update'

        print(request.data)
        request_snapshot = _snapshot_request_data(request.data)
        logger.info("Exam %s request received: %s", action_name, request_snapshot)

        serializer = self.get_serializer(instance=instance, data=request.data, partial=partial)

        try:
            serializer.is_valid(raise_exception=True)
        except DRFValidationError as exc:
            print("Exam serializer errors:", exc.detail)
            logger.warning(
                "Exam %s validation failed: data=%s errors=%s",
                action_name,
                request_snapshot,
                exc.detail,
            )
            raise

        if instance is None:
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

        self.perform_update(serializer)
        if getattr(instance, '_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}
        return Response(serializer.data, status=status.HTTP_200_OK)

    def create(self, request, *args, **kwargs):
        return self._save_exam(request)

    def update(self, request, *args, **kwargs):
        return self._save_exam(request, instance=self.get_object())

    def partial_update(self, request, *args, **kwargs):
        return self._save_exam(request, partial=True, instance=self.get_object())

class SeatingViewSet(viewsets.ModelViewSet):
    queryset = Seating.objects.all().select_related('exam', 'exam_group__program', 'room', 'student')
    serializer_class = SeatingSerializer

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all().select_related('exam', 'room').order_by('-created_at')
    serializer_class = AttendanceSerializer

class LoginView(generics.GenericAPIView):
    permission_classes = [AllowAny]
    serializer_class = UserSerializer

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        if user:
            refresh = RefreshToken.for_user(user)
            return Response({
                'token': str(refresh.access_token),
                'user': UserSerializer(user).data
            })
        return Response({'message': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

class SeedAdminView(generics.GenericAPIView):
    permission_classes = [AllowAny]

    def post(self, request):
        if not User.objects.filter(username='admin').exists():
            User.objects.create_superuser('admin', 'admin@example.com', 'admin123')
            return Response({'message': 'Admin created (admin/admin123)'})
        return Response({'message': 'Admin already exists'})

from django.http import JsonResponse

def chrome_devtools_config(request):
    """
    Handle requests to .well-known/appspecific/com.chrome.devtools.json
    and return an empty JSON response.
    """
    return JsonResponse({})
